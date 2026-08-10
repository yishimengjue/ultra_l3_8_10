# -*- coding: utf-8 -*-
"""生成 L3 四条线封板样例汇总 Excel：总览 sheet + 4 条线各 30 条样例分 sheet。"""
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path("D:/ultra-fineweb-l3-repro/ultra-fineweb-l3-repro")
OUT = ROOT / "L3_delivery/samples/L3_四线封板样例汇总.xlsx"

def load(p, n=None):
    rows = [json.loads(l) for l in open(ROOT / p, encoding="utf-8")]
    return rows

SOURCES = {
    "zh_QA": "QA/claude_model_test/qa_27b_v330_n100_CLEAN_full.jsonl",
    "en_QA": "QA/outputs/qa_en_official500_c16/qa_full.jsonl",
    "zh_rewrite": "rewrite/outputs/rewrite_v4_3_5_qwen3_6_27b_v5090_n500/multi_style_full.jsonl",
    "en_rewrite": "en_rewrite/outputs/en_rewrite_v3_7_l1_n50_enfilter3_2026-08-03/multi_style_full.jsonl",
}

# 封板元信息（真实评测数字）
META = {
    "zh_QA":      dict(prompt="qa_gen_v3_30_forcetype_zh(封板)", closeness="2.76(DMX判官)", quality="忠实4.73 / 可答4.60 / 自足3.65 / 覆盖3.24", note="QA content=原文+问答对(原文在前); CLEAN产出四类脏数据归零; 判官DMX claude-sonnet-4-6与其余线35b口径不同"),
    "en_QA":      dict(prompt="qa_gen_en_v1(hardened)", closeness="3.76", quality="忠实5.0 / 可答5.0 / 自足4.99 / 覆盖3.98", note="QA content=原文+问答对(原文在前)"),
    "zh_rewrite": dict(prompt="rewrite_styles_v4.3.5-v5090", closeness="4.62(style)", quality="忠实4.87 / overall4.74 / 信息保留4.90", note="4风格:encyclopedia/textbook/blog/abstract"),
    "en_rewrite": dict(prompt="rewrite_styles_v3.7", closeness="人工4.99", quality="忠实5.0 / overall4.98 / 人工198/200 pass", note="4风格;历史用阿里云强模型"),
}

wb = openpyxl.Workbook()

# ---------- 总览 sheet ----------
ws = wb.active
ws.title = "总览"
title_fill = PatternFill("solid", fgColor="2F5496")
hdr_fill = PatternFill("solid", fgColor="8EAADB")
thin = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.merge_cells("A1:F1")
c = ws.cell(1, 1, "Ultra-FineWeb-L3 精炼层 · 四条线封板样例汇总")
c.font = Font(bold=True, size=14, color="FFFFFF"); c.fill = title_fill
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

ws.append([])
headers = ["线", "封板 prompt", "接近官方(closeness)", "核心质量(判官1-5)", "样例数", "备注"]
ws.append(headers)
for i in range(1, len(headers)+1):
    cell = ws.cell(3, i); cell.font = Font(bold=True); cell.fill = hdr_fill; cell.border = border
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

line_names = {"zh_QA":"中文 QA","en_QA":"英文 QA","zh_rewrite":"中文 rewrite","en_rewrite":"英文 rewrite"}
for key in ["zh_QA","en_QA","zh_rewrite","en_rewrite"]:
    m = META[key]
    ws.append([line_names[key], m["prompt"], m["closeness"], m["quality"], 30, m["note"]])
    r = ws.max_row
    for i in range(1, len(headers)+1):
        ws.cell(r, i).border = border
        ws.cell(r, i).alignment = Alignment(wrap_text=True, vertical="top")

ws.append([])
note_r = ws.max_row + 1
ws.merge_cells(f"A{note_r}:F{note_r}")
nc = ws.cell(note_r, 1, "口径说明：QA 用 pairwise 对官方问答的 closeness(0-5，越接近官方越高)；rewrite 用 official_style 质量评。两套体系不可直接比大小。判官 qwen3.6-35b-a3b 人工校准过偏宽，读作'无红旗'。所有数字来自真实评测输出，无编造。")
nc.font = Font(italic=True, color="555555"); nc.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[note_r].height = 46

for col, w in zip("ABCDEF", [14, 26, 20, 30, 8, 34]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A4"

# ---------- 各线样例 sheet ----------
win_fill = PatternFill("solid", fgColor="E2EFDA")
def add_line_sheet(key, is_rewrite):
    rows = load(SOURCES[key])
    if is_rewrite:
        # 跨4风格均匀取：每风格取前若干，凑30
        by_style = {}
        for r in rows:
            by_style.setdefault(r.get("_internal_style","?"), []).append(r)
        picked = []
        styles = [s for s in ["encyclopedia","textbook","blog","abstract"] if s in by_style]
        # 轮转取样凑满 30：逐轮从每个风格取一条
        pos = {s: 0 for s in styles}
        while len(picked) < 30:
            advanced = False
            for s in styles:
                if pos[s] < len(by_style[s]) and len(picked) < 30:
                    picked.append(by_style[s][pos[s]]); pos[s] += 1; advanced = True
            if not advanced:
                break
        picked = picked[:30]
    else:
        picked = rows[:30]

    ws = wb.create_sheet(line_names[key])
    m = META[key]
    ws.merge_cells("A1:E1")
    t = ws.cell(1,1, f"{line_names[key]} 封板样例 · prompt={m['prompt']} · closeness={m['closeness']} · {m['quality']}")
    t.font = Font(bold=True, size=11, color="FFFFFF"); t.fill = title_fill
    t.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 34

    if is_rewrite:
        cols = ["#","uid","风格(style)","原文(source)","改写产出(content)"]
    else:
        cols = ["#","uid","官方schema style","原文(source)","QA产出=原文+问答对(content)"]
    ws.append(cols)
    for i in range(1,len(cols)+1):
        cc = ws.cell(2,i); cc.font=Font(bold=True); cc.fill=hdr_fill; cc.border=border
        cc.alignment=Alignment(wrap_text=True, vertical="center", horizontal="center")

    for idx, r in enumerate(picked, 1):
        style_col = r.get("_internal_style") if is_rewrite else r.get("style")
        ws.append([idx, r.get("uid",""), style_col, r.get("_source_text",""), r.get("content","")])
        rr = ws.max_row
        for i in range(1,len(cols)+1):
            ws.cell(rr,i).alignment=Alignment(wrap_text=True, vertical="top")
            ws.cell(rr,i).border=border

    for col, w in zip("ABCDE", [4, 20, 14, 60, 70]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"
    return len(picked)

counts = {}
for key in ["zh_QA","en_QA","zh_rewrite","en_rewrite"]:
    counts[key] = add_line_sheet(key, is_rewrite=("rewrite" in key))

wb.save(OUT)
print("已生成:", OUT)
for k,v in counts.items(): print(f"  {k}: {v} 条样例")
